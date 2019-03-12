import openpyxl
import tkinter 
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import*

wb=openpyxl.load_workbook('register.xlsx')
ws = wb['Sheet1']

window=Tk()

window.geometry("500x900")

label=tkinter.Label(window,text="Register",font=("bold",20))
label.pack()


label_name=tkinter.Label(window,text="Name",font=("bold",14),pady=(20))
label_name.pack()

name=tkinter.Entry(window)
name.pack()

def search():
    result_name.config(text="")
    result_phone.config(text="")
    result_addr.config(text="")
    result_add.config(text="")
    

    for i in range(2,(ws.max_row)+1):
         if((name.get()==ws['A'+str(i)].value) and (phone.get()==ws['B'+str(i)].value)):
             Found=True
             break

         else:
            Found = False
            
            
    if (Found==True):
        result_name.config(text="name: "+str(ws['A'+str(i)].value))
        result_phone.config(text="phone: "+str(ws['B'+str(i)].value))
        result_addr.config(text="address: "+str(ws['C'+str(i)].value))
    
    else:
        result_add.config(text="No record found.")       
          




label_phone=tkinter.Label(window,text="Phone",font=("bold",14),pady=(20))
label_phone.pack()
phone=tkinter.Entry(window)
phone.pack()


label_add=tkinter.Label(window,text="Address",font=("bold",14),pady=(20))
label_add.pack()
add1=tkinter.Entry(window)
add1.pack()


def add():
    for i in range(2,(ws.max_row)+1):
        if((name.get()==ws['A'+str(i)].value) and (phone.get()==ws['B'+str(i)].value)):
            Found=True
            break

        else:
            Found = False

    if(Found==True):
         messagebox.showinfo("Error", "Name Exists!")
                
    else:
        messagebox.showinfo("Success", "Record Added!")
        lastRow=str((ws.max_row)+1)
        ws['A'+lastRow]=name.get()
        ws['B'+lastRow]=phone.get()
        ws['C'+lastRow]=add1.get()

    wb.save('register.xlsx')



bsearch=tkinter.Button(window,text="Search",command=search,font=("Helevetica",11,"bold"))
bsearch.pack()

asearch=tkinter.Button(window,text="Add",command=add,font=("Helevetica",11,"bold"))



def edit():
    for i in range(2,(ws.max_row)+1):
        if((name.get()==ws['A'+str(i)].value) and (phone.get()==ws['B'+str(i)].value)):
            Found=True
            break

        else:
            Found = False

    if(Found==True):
             
            top=Toplevel()
            top.geometry("500x600")

            A=""
            
            def nam():
                global A
                if(var1.get()==1):
                    na=name.get()
                    s.set(na)
                    A=na
                elif(var1.get()==0):
                    s.set("")
            
                
            def ph():
                if(var2.get()==1):
                    ph=phone.get()
                    t.set(ph)
                    B=ph
                elif(var2.get()==0):
                    t.set("")

            def ad():
                if(var3.get()==1):
                    add=add1.get()
                    u.set(add)
                    C=add
                elif(var3.get()==0):
                    u.set("")

                
            label_name=tkinter.Label(top,text="New Name",font=("bold",14),pady=(20))
            label_name.pack()

            s = StringVar()
            t = StringVar()
            u = StringVar()


            
            name1=Entry(top,textvariable=s)
            name1.pack()

            var1 = IntVar()
            c1=Checkbutton(top, text="same as before",variable=var1,command=nam)
            c1.pack()
            
            label_phone=tkinter.Label(top,text="New phone",font=("bold",14),pady=(20))
            label_phone.pack()

            phone1=Entry(top,textvariable=t)
            phone1.pack()

            var2 = IntVar()
            c2=Checkbutton(top, text="same as before", variable=var2,command=ph)
            c2.pack()
                
            label_add=tkinter.Label(top,text="New Address",font=("bold",14),pady=(20))
            label_add.pack()

            add1=Entry(top,textvariable=u)
            add1.pack()

            var3 = IntVar()
            c3=Checkbutton(top, text="same as before", variable=var3,command=ad)
            c3.pack()
            

            def update():
                print(ws['A'+str(i)])
                print(A)


            
            
            Bupdate=tkinter.Button(top,text="Update",command=update,font=("Helevetica",11,"bold"))
            Bupdate.pack()    

            
            top.mainloop()
        
        
                
    else:
        result_add.config(text="no record found to update!")


def delete():
    print()

var1 = IntVar()
Checkbutton(window, text="no change", variable=var1)
edit=tkinter.Button(window,text="Edit",command=edit,font=("Helevetica",11,"bold"))

dsearch=tkinter.Button(window,text="Delete",command=delete,font=("Helevetica",11,"bold"))



asearch.pack()
edit.pack()
dsearch.pack()


res=tkinter.Label(window,text="Result",font=("bold",17,"underline"),pady=20)
res.pack()



result_name=tkinter.Label(window,font=("bold",15))
result_name.pack()

result_phone=tkinter.Label(window,font=("bold",15))
result_phone.pack()

result_addr=tkinter.Label(window,font=("bold",15))
result_addr.pack()

result_add=tkinter.Label(window,font=("bold",15))
result_add.pack()


window.mainloop()


